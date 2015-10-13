"""
Proto type that does the following:

input1:Excel file that contains only numbers.
input2:Excel file that contains the mapping of number to strings.
output 1:Copy of input1, with serial numbers replaced with language A.
output 2:Copy of input1, with serial numbers replaced with language B.
"""

import shutil
from openpyxl import load_workbook, Workbook

shutil.copy('sample-input-fortest-out.xlsx','alter.xlsx')

#Want to save file in two langs, so have two instances.
wb_numbers_1 = load_workbook('sample-input-fortest-out.xlsx')
wb_numbers_2 = load_workbook('alter.xlsx')
wb_translation = load_workbook('translated_file.xlsx')

#Initialize a dictionary that maps a number to a pair of Strings
num_trans_dict = dict()
for sheet in wb_translation:
   for row in sheet.iter_rows():
      triple = []
      for cell in row:
         if cell.value is not None:
            triple.append(cell.value)
      num_trans_dict[triple[0]]=(tuple(triple[1:3]))   
#print num_trans_dict

#go trhough the excel file, extract strings & replace with number.
for sheet in wb_numbers_1:
   for row in sheet.rows:
      for cell in row:
         if None!=cell.value:
            print cell.value
            cell.value=num_trans_dict[int(cell.value)][0]
#save the file containing numbers that replaced the string.
wb_numbers_1.save('product_1.xlsx')

for sheet in wb_numbers_2:
   for row in sheet.rows:
      for cell in row:
         if None!=cell.value:
            print cell.value
            cell.value=num_trans_dict[int(cell.value)][1]
#save the file containing numbers that replaced the string.
wb_numbers_2.save('product_2.xlsx')
"""
#save the extracted strings
wb_out = Workbook(write_only=True) 
ws = wb_out.create_sheet()
for string in string_list:
   ws.append(string)
wb_out.save('new_big_file.xlsx') 
"""
