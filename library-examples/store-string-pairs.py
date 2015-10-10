import sqlite3, string
from openpyxl import load_workbook

#point to the file to be read. Intuitive.
wb = load_workbook('translated_file.xlsx')

str_list = list()
for sheet in wb:
   for row in sheet.iter_rows():
      triple = []
      for cell in row:
         if cell.value is not None:
            triple.append(cell.value)
      str_list.append(tuple(triple))

conn = sqlite3.connect('trans_memory.db')

c = conn.cursor()
for pairs in str_list:
   print pairs
   try:
      c.execute('''INSERT INTO trans_mem
             values(?,?)''',(pairs[1],pairs[2]))
   except sqlite3.IntegrityError:
      print "notification: value already exists."
conn.commit()
conn.close()
