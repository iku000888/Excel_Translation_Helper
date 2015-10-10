from openpyxl import load_workbook, Workbook
#point to the file to be read. Intuitive.
wb2 = load_workbook('sample-input-fortest.xlsx')

#convince your self that sheet names are retireved.
sheet_names = wb2.get_sheet_names()
print sheet_names

#work book is simply a list of sheets
sheet = wb2[sheet_names[0]]
print sheet

print "can iterate sheets, rows and columns intuitively"
string_list = list()
string_list.append(("sequence_number","original language"))
seq_no = 1
for sheet in wb2:
   for row in sheet.rows:
      for cell in row:
         if None!=cell.value:
            string_list.append((seq_no,cell.value))
            seq_no+=1
wb_out = Workbook(write_only=True) 
ws = wb_out.create_sheet()
for string in string_list:
   ws.append(string)

wb_out.save('new_big_file.xlsx') 
